"""
Excel document parser using openpyxl.
"""

from pathlib import Path
from typing import List, Dict, Any

from ragBaseMaker.parsers.base_parser import BaseParser, ParsedDocument


class ExcelParser(BaseParser):
    """Parser for Excel spreadsheets."""
    
    SUPPORTED_EXTENSIONS = ['.xlsx', '.xls', '.xlsm']
    
    def __init__(self, encoding: str = 'utf-8', include_formulas: bool = False):
        super().__init__(encoding)
        self.include_formulas = include_formulas
    
    def parse(self, file_path: str | Path) -> ParsedDocument:
        """
        Parse an Excel document.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            ParsedDocument with extracted text and metadata
        """
        from openpyxl import load_workbook
        
        path = self._validate_file(file_path)
        
        # Load workbook (data_only=True to get values, not formulas)
        wb = load_workbook(path, data_only=not self.include_formulas)
        
        sections: List[Dict[str, Any]] = []
        all_text_parts: List[str] = []
        metadata: Dict[str, Any] = {
            'sheet_count': len(wb.sheetnames),
            'sheet_names': wb.sheetnames,
        }
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Get sheet dimensions
            max_row = sheet.max_row or 0
            max_col = sheet.max_column or 0
            
            if max_row == 0 or max_col == 0:
                continue
            
            sheet_data: List[List[str]] = []
            
            for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                row_data = []
                for cell in row:
                    value = cell.value
                    if value is not None:
                        row_data.append(str(value).strip())
                    else:
                        row_data.append('')
                
                # Only add rows that have at least one non-empty cell
                if any(cell for cell in row_data):
                    sheet_data.append(row_data)
            
            if sheet_data:
                sections.append({
                    'type': 'sheet',
                    'name': sheet_name,
                    'rows': len(sheet_data),
                    'columns': len(sheet_data[0]) if sheet_data else 0,
                    'data': sheet_data,
                })
                
                # Convert to text
                all_text_parts.append(f"\n[Sheet: {sheet_name}]\n")
                for row in sheet_data:
                    all_text_parts.append(' | '.join(row))
        
        wb.close()
        
        content = self._clean_text('\n'.join(all_text_parts))
        
        return ParsedDocument(
            content=content,
            source_path=str(path.absolute()),
            file_type='xlsx',
            metadata=metadata,
            title=path.stem,
            sections=sections,
        )

